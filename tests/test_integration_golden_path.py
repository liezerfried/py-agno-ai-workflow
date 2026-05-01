"""
Integration test against the static golden_input.xlsx fixture.

The file is committed to the repo so any regression in file-reading,
deduplication, or normalization is caught against a concrete, reviewable
artifact rather than dynamically generated data.

LLM is stubbed — the test is deterministic in CI.
"""

from __future__ import annotations

from pathlib import Path
from unittest.mock import MagicMock, patch

import openpyxl
import pytest

from agno.workflow import StepInput

from agents.audit_writer_agent import AuditResult, audit_executor
from agents.ingest_agent import IngestResult, ingest_executor
from agents.mapper_agent import MappingResult, SemanticMatch, mapper_executor
from agents.validator_agent import ValidatorResult, validator_executor
from tests.conftest import PipelineResult, VALID_CATEGORIES, VALID_CATEGORIES_SET
from workflows.pipeline import PipelineError


def run_pipeline(file_path: str, valid_categories: list[str]) -> PipelineResult:
    session = {"file_path": file_path, "target_column": "job_title", "valid_categories": valid_categories}

    ingest_out = ingest_executor(StepInput(), session)
    if not ingest_out.success:
        raise PipelineError("IngestAgent", ingest_out.content)

    validator_out = validator_executor(StepInput(previous_step_content=ingest_out.content), session)
    if not validator_out.success:
        raise PipelineError("ValidatorAgent", validator_out.content)

    mapper_out = mapper_executor(StepInput(previous_step_content=validator_out.content), session)
    if not mapper_out.success:
        raise PipelineError("MapperAgent", mapper_out.content)

    audit_out = audit_executor(StepInput(previous_step_content=mapper_out.content), session)
    if not audit_out.success:
        raise PipelineError("AuditWriter", audit_out.content)

    audit = AuditResult.model_validate_json(audit_out.content)
    return PipelineResult(
        ingest=IngestResult.model_validate_json(ingest_out.content),
        validator=ValidatorResult.model_validate_json(validator_out.content),
        mapper=MappingResult.model_validate_json(mapper_out.content),
        audit=audit,
        output_path=audit.output_path,
    )

FIXTURE_PATH = Path(__file__).parent / "fixtures" / "golden_input.xlsx"

# Row content of golden_input.xlsx — must stay in sync with the file
GOLDEN_ROWS = [
    "Software Developers",           # already valid
    "Lead Accountants and Auditors", # seniority strip → 0.91 → fuzzy auto-correct
    "Fronted Developer",             # typo → LLM band
    "RRHH",                          # abbreviation → LLM
]

_STUB_NO_MATCH = SemanticMatch(
    is_equivalent=False,
    canonical_title=None,
    normalization_type="unknown",
)


def _stub_run(*args, **kwargs) -> MagicMock:
    mock = MagicMock()
    mock.content = _STUB_NO_MATCH
    return mock


def test_golden_input_fixture_exists() -> None:
    assert FIXTURE_PATH.exists(), f"Golden fixture missing: {FIXTURE_PATH}"


def test_golden_path_full_pipeline() -> None:
    """Full pipeline against the static golden fixture with stub LLM."""
    with patch("agents.mapper_agent._mapper_agent") as mock_agent:
        mock_agent.run.side_effect = _stub_run
        result = run_pipeline(str(FIXTURE_PATH), VALID_CATEGORIES)

    unique_rows = len(set(GOLDEN_ROWS))

    # Structural invariants
    assert result.ingest.total_rows == len(GOLDEN_ROWS)
    assert len(result.ingest.raw_categories) == unique_rows
    assert result.validator.valid_count + result.validator.anomaly_count == unique_rows
    assert len(result.mapper.decisions) == result.validator.anomaly_count

    # Hard invariant: no hallucinations
    assert result.audit.hallucination_count == 0
    for d in result.mapper.decisions:
        if d.corrected is not None:
            assert d.corrected in VALID_CATEGORIES_SET

    # "Lead Accountants and Auditors" must be auto-corrected (seniority strip → 0.91 → fuzzy)
    assert result.audit.corrected_count >= 1

    # Output Excel must exist with both required sheets
    assert Path(result.output_path).exists()
    wb = openpyxl.load_workbook(result.output_path)
    assert "Corrected" in wb.sheetnames
    assert "Review Queue" in wb.sheetnames