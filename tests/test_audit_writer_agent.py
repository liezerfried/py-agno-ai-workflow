"""
Unit tests for audit_writer_agent._write_excel.

These verify the Excel output structure and correctness of the three row-outcome
paths (already valid, corrected, needs_review) without going through the executor.
The executor path is covered by test_integration_seams.py::TestMapperToAuditWriter.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from agents.audit_writer_agent import _write_excel
from agents.mapper_agent import MappingDecision
from tests.conftest import VALID_CATEGORIES, VALID_CATEGORIES_SET, make_excel


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _corrected_decision(raw: str, corrected: str) -> MappingDecision:
    return MappingDecision(
        raw=raw,
        preprocessed=raw.lower(),
        corrected=corrected,
        confidence=0.95,
        method="fuzzy",
        normalization_type="typo",
        needs_review=False,
    )


def _review_decision(raw: str, reason: str = "low_confidence") -> MappingDecision:
    return MappingDecision(
        raw=raw,
        preprocessed=raw.lower(),
        corrected=None,
        confidence=0.40,
        method="needs_review",
        normalization_type="unknown",
        needs_review=True,
        review_reason=reason,
    )


def _load_corrected_sheet(output_path: str):
    wb = openpyxl.load_workbook(output_path)
    ws = wb["Corrected"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    return headers, rows


# ---------------------------------------------------------------------------
# Output file structure
# ---------------------------------------------------------------------------


class TestOutputStructure:
    def test_creates_two_sheets(self, tmp_path: Path) -> None:
        xl = make_excel(tmp_path / "in.xlsx", "job_title", ["Software Developers"])
        output_path, *_ = _write_excel(str(xl), {}, "job_title", VALID_CATEGORIES_SET)

        wb = openpyxl.load_workbook(output_path)
        assert "Corrected" in wb.sheetnames
        assert "Review Queue" in wb.sheetnames

    def test_corrected_sheet_has_three_extra_columns(self, tmp_path: Path) -> None:
        xl = make_excel(tmp_path / "in.xlsx", "job_title", ["Software Developers"])
        output_path, *_ = _write_excel(str(xl), {}, "job_title", VALID_CATEGORIES_SET)

        headers, _ = _load_corrected_sheet(output_path)
        assert "corrected_category" in headers
        assert "correction_method" in headers
        assert "confidence" in headers

    def test_review_queue_sheet_columns(self, tmp_path: Path) -> None:
        xl = make_excel(tmp_path / "in.xlsx", "job_title", ["xyz999"])
        decisions = {"xyz999": _review_decision("xyz999")}
        output_path, *_ = _write_excel(str(xl), decisions, "job_title", VALID_CATEGORIES_SET)

        wb = openpyxl.load_workbook(output_path)
        ws = wb["Review Queue"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        assert headers == ["original_category", "preprocessed", "review_reason"]

    def test_output_filename_contains_timestamp(self, tmp_path: Path) -> None:
        xl = make_excel(tmp_path / "myfile.xlsx", "job_title", ["Software Developers"])
        output_path, *_ = _write_excel(str(xl), {}, "job_title", VALID_CATEGORIES_SET)

        stem = Path(output_path).stem
        assert stem.startswith("myfile_corrected_")
        # timestamp part: YYYYMMDD_HHMMSS — 15 chars after prefix
        timestamp_part = stem.replace("myfile_corrected_", "")
        assert len(timestamp_part) == 15
        assert "_" in timestamp_part


# ---------------------------------------------------------------------------
# Row outcome paths
# ---------------------------------------------------------------------------


class TestRowOutcomes:
    def test_already_valid_row_passed_through(self, tmp_path: Path) -> None:
        xl = make_excel(tmp_path / "in.xlsx", "job_title", ["Software Developers"])
        # No decision for "Software Developers" — it was already valid at validator step
        output_path, total_rows, corrected, review, hallucinations = _write_excel(
            str(xl), {}, "job_title", VALID_CATEGORIES_SET
        )

        assert total_rows == 1
        assert corrected == 0
        assert review == 0
        assert hallucinations == 0

        headers, rows = _load_corrected_sheet(output_path)
        corrected_col = headers.index("corrected_category")
        method_col = headers.index("correction_method")
        assert rows[0][corrected_col] == "Software Developers"
        assert rows[0][method_col] == "exact"

    def test_corrected_row_written_with_decision(self, tmp_path: Path) -> None:
        xl = make_excel(tmp_path / "in.xlsx", "job_title", ["Fronted Developer"])
        decisions = {"Fronted Developer": _corrected_decision("Fronted Developer", "Software Developers")}
        output_path, total_rows, corrected, review, hallucinations = _write_excel(
            str(xl), decisions, "job_title", VALID_CATEGORIES_SET
        )

        assert total_rows == 1
        assert corrected == 1
        assert review == 0
        assert hallucinations == 0

        headers, rows = _load_corrected_sheet(output_path)
        corrected_col = headers.index("corrected_category")
        assert rows[0][corrected_col] == "Software Developers"

    def test_needs_review_row_has_blank_correction(self, tmp_path: Path) -> None:
        xl = make_excel(tmp_path / "in.xlsx", "job_title", ["xyz999"])
        decisions = {"xyz999": _review_decision("xyz999")}
        output_path, total_rows, corrected, review, hallucinations = _write_excel(
            str(xl), decisions, "job_title", VALID_CATEGORIES_SET
        )

        assert total_rows == 1
        assert corrected == 0
        assert review == 1
        assert hallucinations == 0

        headers, rows = _load_corrected_sheet(output_path)
        corrected_col = headers.index("corrected_category")
        assert rows[0][corrected_col] is None

    def test_needs_review_row_appears_in_review_sheet(self, tmp_path: Path) -> None:
        xl = make_excel(tmp_path / "in.xlsx", "job_title", ["xyz999"])
        decisions = {"xyz999": _review_decision("xyz999", reason="low_confidence")}
        output_path, *_ = _write_excel(str(xl), decisions, "job_title", VALID_CATEGORIES_SET)

        wb = openpyxl.load_workbook(output_path)
        ws = wb["Review Queue"]
        review_rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert any(r[0] == "xyz999" and r[2] == "low_confidence" for r in review_rows)

    def test_hallucination_rejected_and_counted(self, tmp_path: Path) -> None:
        xl = make_excel(tmp_path / "in.xlsx", "job_title", ["RRHH"])
        # "Chief People Officer" is not in VALID_CATEGORIES_SET — hallucination
        d = MappingDecision(
            raw="RRHH",
            preprocessed="rrhh",
            corrected="Chief People Officer",
            confidence=0.80,
            method="llm",
            normalization_type="unknown",
            needs_review=False,
        )
        output_path, total_rows, corrected, review, hallucinations = _write_excel(
            str(xl), {"RRHH": d}, "job_title", VALID_CATEGORIES_SET
        )

        assert total_rows == 1
        assert hallucinations == 1
        assert corrected == 0
        assert review == 1

        headers, rows = _load_corrected_sheet(output_path)
        corrected_col = headers.index("corrected_category")
        assert rows[0][corrected_col] is None


# ---------------------------------------------------------------------------
# Multi-row and mixed scenarios
# ---------------------------------------------------------------------------


class TestMixedRows:
    def test_original_columns_preserved(self, tmp_path: Path) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["id", "job_title", "department"])
        ws.append([1, "Fronted Developer", "Engineering"])
        path = tmp_path / "multi.xlsx"
        wb.save(path)

        decisions = {"Fronted Developer": _corrected_decision("Fronted Developer", "Software Developers")}
        output_path, *_ = _write_excel(str(path), decisions, "job_title", VALID_CATEGORIES_SET)

        headers, rows = _load_corrected_sheet(output_path)
        assert "id" in headers
        assert "department" in headers
        id_col = headers.index("id")
        dept_col = headers.index("department")
        assert rows[0][id_col] == 1
        assert rows[0][dept_col] == "Engineering"

    def test_mixed_valid_corrected_review(self, tmp_path: Path) -> None:
        xl = make_excel(
            tmp_path / "in.xlsx",
            "job_title",
            ["Software Developers", "Fronted Developer", "xyz999"],
        )
        decisions = {
            "Fronted Developer": _corrected_decision("Fronted Developer", "Software Developers"),
            "xyz999": _review_decision("xyz999"),
        }
        _, total_rows, corrected, review, hallucinations = _write_excel(
            str(xl), decisions, "job_title", VALID_CATEGORIES_SET
        )

        assert total_rows == 3
        assert corrected == 1
        assert review == 1
        assert hallucinations == 0

    def test_row_count_matches_source(self, tmp_path: Path) -> None:
        rows = ["Software Developers", "Data Scientists", "Fronted Developer"]
        xl = make_excel(tmp_path / "in.xlsx", "job_title", rows)
        decisions = {"Fronted Developer": _corrected_decision("Fronted Developer", "Software Developers")}
        output_path, *_ = _write_excel(str(xl), decisions, "job_title", VALID_CATEGORIES_SET)

        _, data_rows = _load_corrected_sheet(output_path)
        assert len(data_rows) == len(rows)