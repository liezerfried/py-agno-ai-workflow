"""
Full end-to-end pipeline integration tests.

Each test runs ingest → validate → map → audit against a real Excel file.
The LLM is stubbed by default.  Mark tests with @pytest.mark.real_llm
and pass --real-llm on the CLI to exercise actual model calls.

Running modes:
    uv run pytest tests/test_integration_pipeline.py              # stub LLM (CI)
    uv run pytest tests/test_integration_pipeline.py --real-llm   # real LLM (integration)
    uv run pytest tests/test_integration_pipeline.py -m "not real_llm"  # explicit stub-only

Parametrize pattern:
    Scenarios are defined as plain dataclasses so they read like a table and
    can be extended without touching test logic.  Each scenario drives the same
    assertion block — adding a new case is two lines.
"""

from __future__ import annotations

import csv
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterator
from unittest.mock import MagicMock, patch

import openpyxl
import pytest

from agno.workflow import StepInput

from agents.audit_writer_agent import AuditResult, audit_executor
from agents.ingest_agent import IngestResult, ingest_executor
from agents.mapper_agent import MappingResult, SemanticMatch, mapper_executor
from agents.validator_agent import ValidatorResult, validator_executor
from tests.conftest import VALID_CATEGORIES, VALID_CATEGORIES_SET, make_excel, make_step_input
from tests.conftest import PipelineResult


# ---------------------------------------------------------------------------
# Pipeline runner (test helper — adds start_from / inject_* for stage isolation)
# ---------------------------------------------------------------------------


def run_pipeline(
    file_path: str,
    target_column: str,
    valid_categories: list[str] | None = None,
    *,
    start_from: str = "ingest",   # "ingest" | "validate" | "map" | "audit"
    inject_ingest: IngestResult | None = None,
    inject_validator: ValidatorResult | None = None,
    inject_mapper: MappingResult | None = None,
) -> PipelineResult:
    """
    Run the normalization pipeline (or a sub-sequence of it).

    start_from controls which stage is the first to execute.  Earlier stages
    can be bypassed by providing pre-built results via inject_*.

    This is the key flexibility mechanism: to test validate+map only, pass
    start_from="validate" and inject_ingest=<your IngestResult>.

    All executor functions are called directly (no Agno Workflow scaffolding)
    so tests are fast and don't depend on async or Chainlit.
    """
    cats = valid_categories if valid_categories is not None else VALID_CATEGORIES

    session: dict = {
        "file_path": file_path,
        "target_column": target_column,
        "valid_categories": cats,
    }

    # Stage 1 — Ingest
    if start_from == "ingest":
        ingest_out = ingest_executor(StepInput(), session)
        assert ingest_out.success, f"IngestAgent failed: {ingest_out.content}"
        ingest_result = IngestResult.model_validate_json(ingest_out.content)
    else:
        assert inject_ingest is not None, "inject_ingest required when start_from != 'ingest'"
        ingest_result = inject_ingest

    ingest_content = ingest_result.model_dump_json()

    # Stage 2 — Validate
    if start_from in ("ingest", "validate"):
        validator_out = validator_executor(make_step_input(ingest_content), session)
        assert validator_out.success, f"ValidatorAgent failed: {validator_out.content}"
        validator_result = ValidatorResult.model_validate_json(validator_out.content)
    else:
        assert inject_validator is not None, "inject_validator required when start_from not in ('ingest','validate')"
        validator_result = inject_validator

    validator_content = validator_result.model_dump_json()

    # Stage 3 — Map
    if start_from in ("ingest", "validate", "map"):
        mapper_out = mapper_executor(make_step_input(validator_content), session)
        assert mapper_out.success, f"MapperAgent failed: {mapper_out.content}"
        mapper_result = MappingResult.model_validate_json(mapper_out.content)
    else:
        assert inject_mapper is not None, "inject_mapper required when start_from == 'audit'"
        mapper_result = inject_mapper

    mapper_content = mapper_result.model_dump_json()

    # Stage 4 — Audit
    audit_out = audit_executor(make_step_input(mapper_content), session)
    assert audit_out.success, f"AuditWriter failed: {audit_out.content}"
    audit_result = AuditResult.model_validate_json(audit_out.content)

    return PipelineResult(
        ingest=ingest_result,
        validator=validator_result,
        mapper=mapper_result,
        audit=audit_result,
        output_path=audit_result.output_path,
    )


# ---------------------------------------------------------------------------
# Scenario definitions
# ---------------------------------------------------------------------------


@dataclass
class Scenario:
    id: str
    rows: list[str]                    # Excel data rows (may repeat)
    target_column: str = "job_title"

    # Expected counts in AuditResult (None = don't assert)
    expected_corrected_min: int = 0    # at least N auto/fuzzy corrections
    expected_review_max: int | None = None  # at most N review-queue entries
    expected_hallucination: int = 0    # must be exactly 0 (default: no hallucinations)
    expected_precision_min: float | None = None  # precision >= threshold (if not None)

    # LLM stub responses in call order (last repeated for remaining calls)
    stub_responses: list[SemanticMatch] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Scenarios designed to exercise specific invariants
# ---------------------------------------------------------------------------

_NO_MATCH = SemanticMatch(
    is_equivalent=False,
    canonical_title=None,
    normalization_type="unknown",
)
_RRHH_MATCH = SemanticMatch(
    is_equivalent=True,
    canonical_title="Human Resources Managers",
    normalization_type="abbreviation",
)

SCENARIOS: list[Scenario] = [
    Scenario(
        id="all_already_valid",
        rows=["Software Developers", "Data Scientists", "Accountants and Auditors"],
        expected_corrected_min=0,
        expected_review_max=0,
        expected_hallucination=0,
        expected_precision_min=None,   # precision=None when nothing attempted
    ),
    Scenario(
        id="high_confidence_typos_only",
        # "Lead Accountants and Auditors" → pre_processor strips "Lead" → "accountants and auditors"
        # → rapidfuzz scores 0.91 (≥ 0.90 threshold) → auto-corrected without LLM.
        # "Accountants and Auditors" is already valid (→ not in anomalies at all).
        rows=["Lead Accountants and Auditors", "Accountants and Auditors"],
        expected_corrected_min=1,      # Lead Accountants auto-corrected
        expected_hallucination=0,
    ),
    Scenario(
        id="low_confidence_all_review",
        rows=["xyz999abc", "qqqqqq", "zzzzz"],
        expected_corrected_min=0,
        expected_review_max=3,
        expected_hallucination=0,
    ),
    Scenario(
        id="mixed_valid_typos_and_junk",
        rows=[
            "Software Developers",        # already valid
            "Lead Accountants and Auditors",  # seniority strip → 0.91 → fuzzy auto-correct
            "RRHH",                       # < 0.70 → needs_review (stub returns no-match anyway)
            "xyzzy_gibberish",            # < 0.70 → review
        ],
        expected_corrected_min=1,      # at least the seniority case is auto-corrected
        expected_hallucination=0,
        stub_responses=[_NO_MATCH],
    ),
    Scenario(
        id="seniority_stripped_then_matched",
        # pre_processor strips seniority → rapidfuzz gets clean tokens.
        # "Lead Accountants and Auditors" → 0.91 (fuzzy auto-correct).
        # "Senior Data Scientists" → 0.87 (LLM band, stub returns no-match → needs_review).
        # Net: at least 1 auto-corrected.
        rows=["Lead Accountants and Auditors", "Senior Data Scientists"],
        expected_corrected_min=1,
        expected_hallucination=0,
        stub_responses=[_NO_MATCH],
    ),
    Scenario(
        id="duplicate_rows_count_correctly",
        # 10 rows but only 2 unique categories: ingest deduplicates.
        # RRHH scores < 0.70 → needs_review without LLM call.
        rows=["Software Developers"] * 7 + ["RRHH"] * 3,
        expected_corrected_min=0,
        expected_hallucination=0,
        stub_responses=[_NO_MATCH],
    ),
    Scenario(
        id="precision_threshold",
        # "Lead Accountants and Auditors" → strip seniority → 0.91 → fuzzy at ≥ 0.90.
        # corrected=1, hallucination=0 → precision=1.0.
        rows=["Lead Accountants and Auditors"],
        expected_corrected_min=1,
        expected_precision_min=1.0,
        expected_hallucination=0,
    ),
]


# ---------------------------------------------------------------------------
# Parametrized full-pipeline tests (stub LLM)
# ---------------------------------------------------------------------------


@pytest.mark.parametrize("scenario", SCENARIOS, ids=[s.id for s in SCENARIOS])
def test_full_pipeline_stub_llm(scenario: Scenario, tmp_path: Path) -> None:
    xl = make_excel(tmp_path / "input.xlsx", scenario.target_column, scenario.rows)

    stub_responses = scenario.stub_responses or [_NO_MATCH]
    call_log: list[SemanticMatch] = []

    def _stub_run(prompt: str, *args, **kwargs):
        idx = min(len(call_log), len(stub_responses) - 1)
        call_log.append(stub_responses[idx])
        mock = MagicMock()
        mock.content = stub_responses[idx]
        return mock

    with patch("agents.mapper_agent._mapper_agent") as mock_agent:
        mock_agent.run.side_effect = _stub_run
        result = run_pipeline(str(xl), scenario.target_column)

    # Structural invariants — always checked regardless of scenario
    total_unique = len(set(scenario.rows))
    assert result.ingest.total_rows == len(scenario.rows)
    assert len(result.ingest.raw_categories) == total_unique
    assert result.validator.valid_count + result.validator.anomaly_count == total_unique
    assert len(result.mapper.decisions) == result.validator.anomaly_count

    # Hard invariant: no invented categories
    for d in result.mapper.decisions:
        if d.corrected is not None:
            assert d.corrected in VALID_CATEGORIES_SET, (
                f"Hallucination: '{d.corrected}' not in valid_categories_set"
            )

    # Scenario-specific assertions
    assert result.audit.hallucination_count == scenario.expected_hallucination
    assert result.audit.corrected_count >= scenario.expected_corrected_min

    if scenario.expected_review_max is not None:
        assert result.audit.review_queue_count <= scenario.expected_review_max

    if scenario.expected_precision_min is not None:
        assert result.audit.precision is not None
        assert result.audit.precision >= scenario.expected_precision_min

    # Output Excel must be a real, readable file with a "Corrected" sheet
    assert Path(result.output_path).exists()
    wb = openpyxl.load_workbook(result.output_path)
    assert "Corrected" in wb.sheetnames
    assert "Review Queue" in wb.sheetnames


# ---------------------------------------------------------------------------
# Stage-pair sub-pipeline tests
# ---------------------------------------------------------------------------


class TestSubPipelines:
    """
    Run individual stage pairs to isolate failures.
    Each method uses run_pipeline(start_from=..., inject_*=...) to skip
    stages that are not the focus.
    """

    def test_validate_map_only(self, tmp_path: Path, stub_llm) -> None:
        """Run only validate → map, injecting a hand-crafted IngestResult."""
        xl = make_excel(tmp_path / "i.xlsx", "job_title", ["Fronted Developer"])
        injected = IngestResult(
            file_path=str(xl),
            target_column="job_title",
            raw_categories=["Fronted Developer"],
            total_rows=1,
        )
        result = run_pipeline(
            str(xl),
            "job_title",
            start_from="validate",
            inject_ingest=injected,
        )
        assert result.validator.anomaly_count >= 1
        assert len(result.mapper.decisions) == result.validator.anomaly_count

    def test_map_audit_only(self, tmp_path: Path, stub_llm) -> None:
        """Run only map → audit, injecting a hand-crafted ValidatorResult."""
        from infrastructure.pipeline.contracts import CategoryValidation

        xl = make_excel(tmp_path / "i.xlsx", "job_title", ["RRHH"])
        injected_ingest = IngestResult(
            file_path=str(xl),
            target_column="job_title",
            raw_categories=["RRHH"],
            total_rows=1,
        )
        injected_validator = ValidatorResult(
            validations=[
                CategoryValidation(raw="RRHH", is_valid=False, closest_match=None, similarity_score=0.55)
            ],
            valid_count=0,
            anomaly_count=1,
            anomalies=[
                CategoryValidation(raw="RRHH", is_valid=False, closest_match=None, similarity_score=0.55)
            ],
        )
        result = run_pipeline(
            str(xl),
            "job_title",
            start_from="map",
            inject_ingest=injected_ingest,
            inject_validator=injected_validator,
        )
        assert len(result.mapper.decisions) == 1


# ---------------------------------------------------------------------------
# Golden dataset — precision / hallucination_rate regression
# ---------------------------------------------------------------------------


@dataclass
class GoldenRow:
    raw: str
    expected_corrected: str | None   # None means needs_review is acceptable


GOLDEN_DATASET: list[GoldenRow] = [
    GoldenRow("Software Developers", "Software Developers"),                 # already valid
    GoldenRow("Softwarre Developers", "Software Developers"),                # typo — 0.87, LLM band
    GoldenRow("Senior Data Scientists", "Data Scientists"),                  # seniority — 0.87, LLM band
    GoldenRow("RRHH", "Human Resources Managers"),                           # abbreviation — LLM (stub: targeted)
    GoldenRow("Desarrollador Backend", None),                                # language — LLM (stub: no-match)
    GoldenRow("xyzzy_gibberish", None),                                      # < 0.70 → review
    GoldenRow("Fronted Developer", "Frontend Web Developers"),               # typo — 0.75, LLM band
    GoldenRow("Lead Accountants and Auditors", "Accountants and Auditors"),  # seniority strip → 0.91 → fuzzy
]

# Evaluation thresholds (from CLAUDE.md)
PRECISION_THRESHOLD = 0.85
HALLUCINATION_RATE_THRESHOLD = 0.05


def test_golden_dataset_metrics(tmp_path: Path) -> None:
    """
    Run the full pipeline against the golden dataset and verify that
    precision ≥ 0.85 and hallucination_rate ≤ 0.05.

    This test is the regression gate — if a pipeline change drops precision
    below the threshold, this test fails and blocks the merge.

    The LLM is stubbed with a targeted response for RRHH so the golden
    dataset remains deterministic.
    """
    raws = [g.raw for g in GOLDEN_DATASET]
    xl = make_excel(tmp_path / "golden.xlsx", "job_title", raws)

    rrhh_response = SemanticMatch(
        is_equivalent=True,
        canonical_title="Human Resources Managers",
        normalization_type="abbreviation",
    )
    fallback = SemanticMatch(
        is_equivalent=False,
        canonical_title=None,
        normalization_type="unknown",
    )

    def targeted_stub(prompt: str, *args, **kwargs):
        mock = MagicMock()
        mock.content = rrhh_response if "RRHH" in prompt else fallback
        return mock

    with patch("agents.mapper_agent._mapper_agent") as mock_agent:
        mock_agent.run.side_effect = targeted_stub
        result = run_pipeline(str(xl), "job_title")

    total_attempted = result.audit.corrected_count + result.audit.hallucination_count
    hallucination_rate = (
        result.audit.hallucination_count / total_attempted if total_attempted > 0 else 0.0
    )

    # Precision check
    if total_attempted > 0:
        assert result.audit.precision >= PRECISION_THRESHOLD, (
            f"Precision {result.audit.precision:.2%} below threshold {PRECISION_THRESHOLD:.0%}"
        )

    # Hallucination rate check
    assert hallucination_rate <= HALLUCINATION_RATE_THRESHOLD, (
        f"Hallucination rate {hallucination_rate:.2%} exceeds threshold {HALLUCINATION_RATE_THRESHOLD:.0%}"
    )

    # At minimum, the fuzzy-auto case (Lead Accountants) must be corrected.
    # LLM-band cases (Softwarre, Senior Data Scientists, RRHH, Fronted Developer)
    # depend on stub responses — their corrections are accounted for separately.
    assert result.audit.corrected_count >= 1, (
        "Expected at least 1 auto-correction from golden dataset "
        f"(got {result.audit.corrected_count})"
    )


# ---------------------------------------------------------------------------
# Real LLM tests (skipped in CI)
# ---------------------------------------------------------------------------


@pytest.mark.real_llm
def test_full_pipeline_real_llm(tmp_path: Path, request: pytest.FixtureRequest) -> None:
    """
    Run the full pipeline with a live LLM call.  Skipped unless --real-llm is passed.

    This test does NOT stub anything — it exercises the actual model configured
    in infrastructure/llm/provider.py (LM Studio or Groq depending on LLM_PROVIDER).
    """
    if not request.config.getoption("--real-llm"):
        pytest.skip("Pass --real-llm to run real LLM tests")

    xl = make_excel(
        tmp_path / "real.xlsx",
        "job_title",
        ["RRHH", "Desarrollador Backend", "Fronted Developer", "Software Developers"],
    )

    result = run_pipeline(str(xl), "job_title")

    # Structural invariants still hold
    for d in result.mapper.decisions:
        if d.corrected is not None:
            assert d.corrected in VALID_CATEGORIES_SET, (
                f"Real LLM hallucinated: '{d.corrected}'"
            )

    # Precision must meet threshold even with real LLM
    total_attempted = result.audit.corrected_count + result.audit.hallucination_count
    if total_attempted > 0:
        assert result.audit.precision >= PRECISION_THRESHOLD


# ---------------------------------------------------------------------------
# Performance / batch size tests
# ---------------------------------------------------------------------------


@pytest.mark.parametrize("batch_size", [10, 100, 500])
def test_pipeline_batch_size(batch_size: int, tmp_path: Path, stub_llm) -> None:
    """
    Verify that the pipeline handles batches without crashing or producing
    structurally invalid output.  Does not assert correctness — only that the
    pipeline completes and all count invariants hold.
    """
    # Mix: half valid, half typos
    half = batch_size // 2
    rows = ["Software Developers"] * half + ["Softwarre Developers"] * (batch_size - half)

    xl = make_excel(tmp_path / f"batch_{batch_size}.xlsx", "job_title", rows)
    result = run_pipeline(str(xl), "job_title")

    assert result.ingest.total_rows == batch_size
    assert len(result.mapper.decisions) > 0
    assert result.audit.hallucination_count == 0


# ---------------------------------------------------------------------------
# 100-row real-data metrics test (requires --real-llm + LM Studio running)
# ---------------------------------------------------------------------------

TEST_100_PATH = Path(__file__).parent.parent / "data" / "test_100_autodetect.xlsx"


@pytest.mark.real_llm
def test_100_anomalies_real_llm_metrics(request: pytest.FixtureRequest) -> None:
    """
    Run the full pipeline against data/test_100_autodetect.xlsx with a live LLM.

    Acceptance criteria (from CLAUDE.md):
      precision          >= 0.85
      hallucination_rate <= 0.05

    Run with:
      uv run pytest tests/test_integration_pipeline.py::test_100_anomalies_real_llm_metrics --real-llm -v -s
    """
    if not request.config.getoption("--real-llm"):
        pytest.skip("Pass --real-llm to run real LLM tests")

    if not TEST_100_PATH.exists():
        pytest.skip(f"Test file not found: {TEST_100_PATH}")

    result = run_pipeline(str(TEST_100_PATH), "job_category")

    total_attempted = result.audit.corrected_count + result.audit.hallucination_count
    hallucination_rate = (
        result.audit.hallucination_count / total_attempted if total_attempted > 0 else 0.0
    )

    # Print breakdown so the run is auditable even when the test passes
    print(f"\n--- 100-row real LLM metrics ---")
    print(f"  Total rows        : {result.ingest.total_rows}")
    print(f"  Unique categories : {len(result.ingest.raw_categories)}")
    print(f"  Anomalies flagged : {result.validator.anomaly_count}")
    print(f"  Already valid     : {result.validator.valid_count}")
    print(f"  Corrected         : {result.audit.corrected_count}")
    print(f"  Review queue      : {result.audit.review_queue_count}")
    print(f"  Hallucinations    : {result.audit.hallucination_count}")
    precision_str = f"{result.audit.precision:.2%}" if result.audit.precision is not None else "N/A"
    print(f"  Precision         : {precision_str}")
    print(f"  Hallucination rate: {hallucination_rate:.2%}")
    print(f"  Output file       : {result.audit.output_path}")

    # Structural invariant: no hallucinated title anywhere in decisions
    for d in result.mapper.decisions:
        if d.corrected is not None:
            assert d.corrected in VALID_CATEGORIES_SET, (
                f"Hallucination: '{d.original}' -> '{d.corrected}' is not in O*NET"
            )

    if total_attempted > 0:
        assert result.audit.precision >= PRECISION_THRESHOLD, (
            f"Precision {result.audit.precision:.2%} below threshold {PRECISION_THRESHOLD:.0%}"
        )

    assert hallucination_rate <= HALLUCINATION_RATE_THRESHOLD, (
        f"Hallucination rate {hallucination_rate:.2%} exceeds threshold {HALLUCINATION_RATE_THRESHOLD:.0%}"
    )