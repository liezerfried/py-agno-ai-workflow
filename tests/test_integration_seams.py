"""
Integration tests for the JSON-serialization seams between adjacent pipeline stages.

Each test class covers exactly one stage boundary:

    Class                       Boundary tested
    ─────────────────────────   ──────────────────────────────
    TestIngestToValidator       ingest_executor → validator_executor
    TestValidatorToMapper       validator_executor → mapper_executor
    TestMapperToAuditWriter     mapper_executor → audit_executor

Why seams and not just unit tests?
- The handshake between stages is a raw JSON string written by model_dump_json()
  and consumed by model_validate_json().  Any field rename, type change, or
  accidental None breaks the next stage silently (caught by the generic except
  and returned as success=False with a cryptic message).
- session_state key names are not typed — a single typo is invisible until runtime.
"""

from __future__ import annotations

from pathlib import Path

import pytest

from agno.workflow import StepInput, StepOutput

from agents.audit_writer_agent import AuditResult, audit_executor
from agents.ingest_agent import IngestResult, ingest_executor
from agents.mapper_agent import MappingDecision, MappingResult, mapper_executor
from agents.validator_agent import ValidatorResult, validator_executor
from tests.conftest import VALID_CATEGORIES, VALID_CATEGORIES_SET, make_excel, make_step_input


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def assert_step_ok(out: StepOutput, label: str = "") -> None:
    """Fail with a descriptive message if a StepOutput is not successful."""
    prefix = f"[{label}] " if label else ""
    assert out.success, f"{prefix}StepOutput.success=False — content: {out.content!r}"


def _base_session(file_path: str = "dummy.xlsx", column: str = "job_title") -> dict:
    return {
        "file_path": file_path,
        "target_column": column,
        "valid_categories": VALID_CATEGORIES,
    }


# ---------------------------------------------------------------------------
# Seam 1 — Ingest → Validator
# ---------------------------------------------------------------------------


class TestIngestToValidator:
    """
    Verifies that whatever ingest_executor writes into StepOutput.content is
    exactly what validator_executor can consume via IngestResult.model_validate_json.
    """

    def test_valid_file_passes_through(self, tmp_path: Path) -> None:
        xl = make_excel(tmp_path / "s.xlsx", "job_title", ["Software Developers", "RRHH"])
        session = _base_session(str(xl), "job_title")

        ingest_out = ingest_executor(StepInput(), session)
        assert_step_ok(ingest_out, "ingest")

        # The output must be valid JSON that deserializes to IngestResult
        parsed = IngestResult.model_validate_json(ingest_out.content)
        assert len(parsed.raw_categories) == 2
        assert parsed.total_rows == 2

        # Pass the raw string (not a re-serialized copy) directly to the next stage
        validator_out = validator_executor(make_step_input(ingest_out.content), session)
        assert_step_ok(validator_out, "validate")

        validator_result = ValidatorResult.model_validate_json(validator_out.content)
        assert validator_result.valid_count + validator_result.anomaly_count == 2

    def test_duplicate_rows_deduplicated(self, tmp_path: Path) -> None:
        xl = make_excel(
            tmp_path / "dup.xlsx",
            "job_title",
            ["Software Developers", "Software Developers", "RRHH"],
        )
        session = _base_session(str(xl), "job_title")

        ingest_out = ingest_executor(StepInput(), session)
        assert_step_ok(ingest_out, "ingest")

        parsed = IngestResult.model_validate_json(ingest_out.content)
        # IngestResult deduplicates — 2 unique values, not 3
        assert len(parsed.raw_categories) == 2

    def test_missing_column_propagates_failure(self, tmp_path: Path) -> None:
        xl = make_excel(tmp_path / "bad.xlsx", "other_col", ["Software Developers"])
        session = _base_session(str(xl), "job_title")  # wrong column name

        ingest_out = ingest_executor(StepInput(), session)
        assert ingest_out.success is False
        assert ingest_out.stop is True

    def test_session_state_key_names(self, tmp_path: Path) -> None:
        """All executors require the full PipelineSession (all 4 keys).
        A missing key causes a KeyError at from_dict() → success=False."""
        xl = make_excel(tmp_path / "k.xlsx", "role", ["Data Scientists"])
        full_session = _base_session(str(xl), "role")

        ingest_out = ingest_executor(StepInput(), full_session)
        assert_step_ok(ingest_out, "ingest key names")

        validator_out = validator_executor(make_step_input(ingest_out.content), full_session)
        assert_step_ok(validator_out, "validator key names")

        # Partial session (missing valid_categories) must fail gracefully
        broken = {"file_path": str(xl), "target_column": "role"}
        broken_out = validator_executor(make_step_input(ingest_out.content), broken)
        assert broken_out.success is False


# ---------------------------------------------------------------------------
# Seam 2 — Validator → Mapper
# ---------------------------------------------------------------------------


class TestValidatorToMapper:
    """
    Verifies that the ValidatorResult JSON that validator_executor emits is
    correctly consumed by mapper_executor, including the anomalies subset.
    """

    def _validator_out_for(self, categories: list[str]) -> StepOutput:
        ingest_result = IngestResult(
            file_path="dummy.xlsx",
            target_column="job_title",
            raw_categories=categories,
            total_rows=len(categories),
        )
        session = _base_session()
        return validator_executor(make_step_input(ingest_result.model_dump_json()), session)

    def test_all_valid_produces_empty_anomalies(self, stub_llm) -> None:
        validator_out = self._validator_out_for(["Software Developers", "Data Scientists"])
        assert_step_ok(validator_out, "validator")

        session = _base_session()
        mapper_out = mapper_executor(make_step_input(validator_out.content), session)
        assert_step_ok(mapper_out, "mapper")

        result = MappingResult.model_validate_json(mapper_out.content)
        assert result.decisions == []

    def test_anomaly_count_matches_decisions_length(self, stub_llm) -> None:
        # Mix of valid and anomalous
        categories = ["Software Developers", "RRHH", "Fronted Developer", "xyz999"]
        validator_out = self._validator_out_for(categories)
        assert_step_ok(validator_out, "validator")

        vr = ValidatorResult.model_validate_json(validator_out.content)

        session = _base_session()
        mapper_out = mapper_executor(make_step_input(validator_out.content), session)
        assert_step_ok(mapper_out, "mapper")

        mr = MappingResult.model_validate_json(mapper_out.content)
        # mapper only processes anomalies — decisions list must have same length
        assert len(mr.decisions) == vr.anomaly_count

    def test_corrected_titles_are_in_valid_set(self, stub_llm) -> None:
        """Hard invariant: every non-None corrected value must be a valid O*NET title."""
        validator_out = self._validator_out_for(["Fronted Developer", "softwaree developer"])
        session = _base_session()
        mapper_out = mapper_executor(make_step_input(validator_out.content), session)
        assert_step_ok(mapper_out, "mapper")

        mr = MappingResult.model_validate_json(mapper_out.content)
        for d in mr.decisions:
            if d.corrected is not None:
                assert d.corrected in VALID_CATEGORIES_SET, (
                    f"Hallucination detected: '{d.corrected}' is not in valid_categories_set"
                )

    def test_missing_valid_categories_key_fails_gracefully(self) -> None:
        """If session_state is missing valid_categories, mapper must return success=False."""
        validator_out = self._validator_out_for(["Fronted Developer"])
        broken_session = {"file_path": "dummy.xlsx", "target_column": "job_title"}  # missing valid_categories

        mapper_out = mapper_executor(make_step_input(validator_out.content), broken_session)
        assert mapper_out.success is False


# ---------------------------------------------------------------------------
# Seam 3 — Mapper → AuditWriter
# ---------------------------------------------------------------------------


class TestMapperToAuditWriter:
    """
    Verifies that MappingResult JSON produced by mapper_executor is correctly
    consumed by audit_executor and that the written Excel matches the decisions.
    """

    def _make_mapping_result(
        self,
        decisions: list[MappingDecision],
    ) -> str:
        """Serialize a MappingResult directly, bypassing earlier stages."""
        return MappingResult(decisions=decisions).model_dump_json()

    def test_corrected_rows_appear_in_excel(self, tmp_path: Path) -> None:
        xl = make_excel(tmp_path / "in.xlsx", "job_title", ["Fronted Developer", "Data Scientists"])

        decisions = [
            MappingDecision(
                raw="Fronted Developer",
                preprocessed="fronted developer",
                corrected="Software Developers",
                confidence=0.95,
                method="fuzzy",
                normalization_type="typo",
                needs_review=False,
            )
            # "Data Scientists" was already valid → not in decisions
        ]
        session = {
            "file_path": str(xl),
            "target_column": "job_title",
            "valid_categories": VALID_CATEGORIES,
        }
        audit_out = audit_executor(
            make_step_input(self._make_mapping_result(decisions)), session
        )
        assert_step_ok(audit_out, "audit")

        ar = AuditResult.model_validate_json(audit_out.content)
        assert ar.corrected_count == 1
        assert ar.review_queue_count == 0
        assert ar.hallucination_count == 0
        assert ar.precision == 1.0

        # Verify the physical Excel was written and contains the correction
        import openpyxl as oxl
        wb = oxl.load_workbook(ar.output_path)
        ws = wb["Corrected"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        assert "corrected_category" in headers
        corrected_col = headers.index("corrected_category")

        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        corrections = [r[corrected_col] for r in data_rows if r[corrected_col] is not None]
        assert "Software Developers" in corrections

    def test_needs_review_rows_go_to_review_sheet(self, tmp_path: Path) -> None:
        xl = make_excel(tmp_path / "in.xlsx", "job_title", ["RRHH"])

        decisions = [
            MappingDecision(
                raw="RRHH",
                preprocessed="rrhh",
                corrected=None,
                confidence=0.55,
                method="needs_review",
                normalization_type="unknown",
                needs_review=True,
                review_reason="low_confidence",
            )
        ]
        session = {
            "file_path": str(xl),
            "target_column": "job_title",
            "valid_categories": VALID_CATEGORIES,
        }
        audit_out = audit_executor(
            make_step_input(self._make_mapping_result(decisions)), session
        )
        assert_step_ok(audit_out, "audit")

        ar = AuditResult.model_validate_json(audit_out.content)
        assert ar.corrected_count == 0
        assert ar.review_queue_count == 1
        assert ar.hallucination_count == 0
        assert ar.precision is None  # no corrections attempted

        import openpyxl as oxl
        wb = oxl.load_workbook(ar.output_path)
        assert "Review Queue" in wb.sheetnames
        ws_review = wb["Review Queue"]
        review_rows = list(ws_review.iter_rows(min_row=2, values_only=True))
        assert any(r[0] == "RRHH" for r in review_rows)

    def test_hallucination_rejected_increments_counter(self, tmp_path: Path) -> None:
        """A corrected title not in valid_categories_set must be counted as hallucination."""
        xl = make_excel(tmp_path / "in.xlsx", "job_title", ["RRHH"])

        decisions = [
            MappingDecision(
                raw="RRHH",
                preprocessed="rrhh",
                corrected="Chief People Officer",  # not in VALID_CATEGORIES_SET
                confidence=0.80,
                method="llm",
                normalization_type="unknown",
                needs_review=False,
            )
        ]
        session = {
            "file_path": str(xl),
            "target_column": "job_title",
            "valid_categories": VALID_CATEGORIES,
        }
        audit_out = audit_executor(
            make_step_input(self._make_mapping_result(decisions)), session
        )
        assert_step_ok(audit_out, "audit")

        ar = AuditResult.model_validate_json(audit_out.content)
        assert ar.hallucination_count == 1
        assert ar.corrected_count == 0
        # Row must appear in review queue, not in corrected column
        assert ar.review_queue_count == 1

    def test_precision_formula(self, tmp_path: Path) -> None:
        """precision = corrected / (corrected + hallucination_count)."""
        xl = make_excel(
            tmp_path / "in.xlsx",
            "job_title",
            ["Fronted Developer", "RRHH", "xyz999"],
        )
        decisions = [
            MappingDecision(
                raw="Fronted Developer",
                preprocessed="fronted developer",
                corrected="Software Developers",  # valid → counted
                confidence=0.95,
                method="fuzzy",
                normalization_type="typo",
                needs_review=False,
            ),
            MappingDecision(
                raw="RRHH",
                preprocessed="rrhh",
                corrected="Chief People Officer",  # hallucination
                confidence=0.75,
                method="llm",
                normalization_type="unknown",
                needs_review=False,
            ),
            MappingDecision(
                raw="xyz999",
                preprocessed="xyz999",
                corrected=None,
                confidence=0.10,
                method="needs_review",
                normalization_type="unknown",
                needs_review=True,
                review_reason="low_confidence",
            ),
        ]
        session = {
            "file_path": str(xl),
            "target_column": "job_title",
            "valid_categories": VALID_CATEGORIES,
        }
        audit_out = audit_executor(
            make_step_input(self._make_mapping_result(decisions)), session
        )
        assert_step_ok(audit_out, "audit")

        ar = AuditResult.model_validate_json(audit_out.content)
        assert ar.corrected_count == 1
        assert ar.hallucination_count == 1
        # precision = 1 / (1 + 1) = 0.5
        assert ar.precision == pytest.approx(0.5)

    def test_raw_key_whitespace_mismatch(self, tmp_path: Path) -> None:
        """
        decisions_by_raw is keyed on MappingDecision.raw.
        The Excel reader does str(cell).strip() when building the lookup key.
        This test ensures a cell value with leading/trailing spaces still matches.
        """
        xl = make_excel(tmp_path / "ws.xlsx", "job_title", ["  Fronted Developer  "])

        decisions = [
            MappingDecision(
                raw="Fronted Developer",  # stripped key as produced by ingest_executor
                preprocessed="fronted developer",
                corrected="Software Developers",
                confidence=0.95,
                method="fuzzy",
                normalization_type="typo",
                needs_review=False,
            )
        ]
        session = {
            "file_path": str(xl),
            "target_column": "job_title",
            "valid_categories": VALID_CATEGORIES,
        }
        audit_out = audit_executor(
            make_step_input(self._make_mapping_result(decisions)), session
        )
        assert_step_ok(audit_out, "audit")
        ar = AuditResult.model_validate_json(audit_out.content)
        # The correction must be applied, not missed due to whitespace in the cell
        assert ar.corrected_count == 1