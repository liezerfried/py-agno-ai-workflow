"""
Unit tests for ingest_agent — scan_headers, detect_job_column, extract_categories.

These cover the module in isolation: no executor, no session state, no pipeline.
The integration seam tests (test_integration_seams.py) verify the executor path.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from agents.ingest_agent import (
    IngestResult,
    detect_job_column,
    extract_categories,
    scan_headers,
)
from tests.conftest import make_excel


# ---------------------------------------------------------------------------
# scan_headers
# ---------------------------------------------------------------------------


class TestScanHeaders:
    def test_returns_all_column_names(self, tmp_path: Path) -> None:
        xl = make_excel(tmp_path / "f.xlsx", "job_title", ["Software Developers"])
        result = scan_headers(str(xl))
        assert result.column_names == ["job_title"]
        assert result.file_path == str(xl)

    def test_multi_column_file(self, tmp_path: Path) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["id", "name", "job_title", "department"])
        ws.append([1, "Alice", "Software Developers", "Engineering"])
        path = tmp_path / "multi.xlsx"
        wb.save(path)

        result = scan_headers(str(path))
        assert result.column_names == ["id", "name", "job_title", "department"]

    def test_filters_none_columns(self, tmp_path: Path) -> None:
        # openpyxl can return None for trailing phantom cells in some spreadsheets
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["job_title", None, "department"])
        path = tmp_path / "none_col.xlsx"
        wb.save(path)

        result = scan_headers(str(path))
        assert None not in result.column_names
        assert "job_title" in result.column_names


# ---------------------------------------------------------------------------
# detect_job_column
# ---------------------------------------------------------------------------


class TestDetectJobColumn:
    def test_exact_english_match(self) -> None:
        col, score = detect_job_column(["job_title", "department", "salary"])
        assert col == "job_title"
        assert score > 0.8

    def test_spanish_column_detected(self) -> None:
        col, score = detect_job_column(["nombre", "cargo", "departamento"])
        assert col == "cargo"
        assert score > 0.8

    def test_empty_list_returns_none(self) -> None:
        col, score = detect_job_column([])
        assert col is None
        assert score == 0.0

    def test_best_match_wins_over_weaker(self) -> None:
        # "job_title" should beat "department"
        col, _ = detect_job_column(["department", "job_title"])
        assert col == "job_title"

    def test_score_normalized_to_0_1(self) -> None:
        _, score = detect_job_column(["job_title"])
        assert 0.0 <= score <= 1.0

    @pytest.mark.parametrize("header", [
        "puesto",
        "rol",
        "categoria",
        "ocupacion",
        "cargo",
        "position",
        "occupation",
        "title",
        "role",
    ])
    def test_bilingual_synonyms_detected(self, header: str) -> None:
        col, score = detect_job_column([header])
        assert col == header
        assert score > 0.7


# ---------------------------------------------------------------------------
# extract_categories
# ---------------------------------------------------------------------------


class TestExtractCategories:
    def test_basic_extraction(self, tmp_path: Path) -> None:
        xl = make_excel(tmp_path / "f.xlsx", "job_title", ["Software Developers", "Data Scientists"])
        result = extract_categories(str(xl), "job_title")

        assert isinstance(result, IngestResult)
        assert result.target_column == "job_title"
        assert result.total_rows == 2
        assert sorted(result.raw_categories) == ["Data Scientists", "Software Developers"]

    def test_deduplicates_values(self, tmp_path: Path) -> None:
        xl = make_excel(
            tmp_path / "dup.xlsx",
            "job_title",
            ["Software Developers", "Software Developers", "Data Scientists"],
        )
        result = extract_categories(str(xl), "job_title")

        assert result.total_rows == 3
        assert len(result.raw_categories) == 2

    def test_output_is_sorted(self, tmp_path: Path) -> None:
        xl = make_excel(tmp_path / "f.xlsx", "job_title", ["Zzz Role", "Aaa Role", "Mmm Role"])
        result = extract_categories(str(xl), "job_title")

        assert result.raw_categories == sorted(result.raw_categories)

    def test_blank_cells_excluded_from_categories(self, tmp_path: Path) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["job_title"])
        ws.append(["Software Developers"])
        ws.append([None])           # blank cell
        ws.append(["Data Scientists"])
        path = tmp_path / "blanks.xlsx"
        wb.save(path)

        result = extract_categories(str(path), "job_title")
        assert result.total_rows == 3          # blank row still counted
        assert len(result.raw_categories) == 2  # blank excluded from unique list
        assert None not in result.raw_categories

    def test_whitespace_stripped(self, tmp_path: Path) -> None:
        xl = make_excel(tmp_path / "ws.xlsx", "job_title", ["  Software Developers  "])
        result = extract_categories(str(xl), "job_title")

        assert "Software Developers" in result.raw_categories
        assert "  Software Developers  " not in result.raw_categories

    def test_missing_column_raises_value_error(self, tmp_path: Path) -> None:
        xl = make_excel(tmp_path / "f.xlsx", "other_column", ["Software Developers"])

        with pytest.raises(ValueError, match="not found"):
            extract_categories(str(xl), "job_title")

    def test_multi_column_file_reads_correct_column(self, tmp_path: Path) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["id", "name", "job_title"])
        ws.append([1, "Alice", "Software Developers"])
        ws.append([2, "Bob", "Data Scientists"])
        path = tmp_path / "multi.xlsx"
        wb.save(path)

        result = extract_categories(str(path), "job_title")
        assert set(result.raw_categories) == {"Software Developers", "Data Scientists"}
        assert "Alice" not in result.raw_categories
        assert "Bob" not in result.raw_categories