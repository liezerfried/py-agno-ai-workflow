"""
Generate two test Excel files:
  - data/test_100_autodetect.xlsx  : column "job_category"  -> rapidfuzz score 1.0, auto-detect
  - data/test_100_ambiguous.xlsx   : column "headcount"     -> score 0.63, shows buttons

All anomaly values are variants of titles that exist in data/valid_categories.csv (O*NET).
Canonical target for each row is noted inline.

Run with: uv run python scripts/generate_test_files.py
"""
import csv

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill


def _load_valid() -> set[str]:
    with open("data/valid_categories.csv") as f:
        return {r[0] for r in csv.reader(f) if r and r[0] != "category"}


# 100 rows covering all 7 normalization types.
# Format: (employee_id, name, department, job_category_anomaly)
# Canonical O*NET target shown in comment on each row.
ROWS = [
    # ── Type 1: Typo — corrected by rapidfuzz ────────────────────────────────
    ("E001", "Ana Garcia",         "Engineering",  "Sofware Developers"),                                    # -> Software Developers
    ("E002", "Luis Ramos",         "QA",           "Sofware Quality Assurence Analysts and Testers"),        # -> Software Quality Assurance Analysts and Testers
    ("E003", "Maria Lopez",        "Finance",      "Finacial and Investement Analysts"),                     # -> Financial and Investment Analysts
    ("E004", "Carlos Perez",       "IT",           "Databse Administrators"),                                # -> Database Administrators
    ("E005", "Sara Jimenez",       "Data",         "Data Sceintists"),                                       # -> Data Scientists
    ("E006", "Pedro Morales",      "Network",      "Netwrok and Computer Systems Administrators"),           # -> Network and Computer Systems Administrators
    ("E007", "Elena Torres",       "Security",     "Informaton Security Analysts"),                          # -> Information Security Analysts
    ("E008", "Miguel Castro",      "Marketing",    "Marketting Managers"),                                   # -> Marketing Managers
    ("E009", "Valentina Diaz",     "Ops",          "Operatons Research Analysts"),                           # -> Operations Research Analysts
    ("E010", "Andres Ruiz",        "Web",          "Web Devlopers"),                                         # -> Web Developers
    ("E011", "Natalia Vega",       "Support",      "Custoer Service Representatives"),                      # -> Customer Service Representatives
    ("E012", "Tomas Herrera",      "IT",           "Compter Systems Analysts"),                              # -> Computer Systems Analysts
    ("E013", "Sofia Medina",       "Nursing",      "Registerd Nurses"),                                      # -> Registered Nurses
    ("E014", "Ricardo Soto",       "Finance",      "Acocuntants and Auditors"),                              # -> Accountants and Auditors
    ("E015", "Julia Romero",       "Engineering",  "Sofware Develoeprs"),                                    # -> Software Developers
    # ── Type 2: Casing / punctuation — corrected by pre_processor ────────────
    ("E016", "Fernando Rios",      "Engineering",  "SOFTWARE DEVELOPERS"),                                   # -> Software Developers
    ("E017", "Gabriela Nunez",     "QA",           "SOFTWARE QUALITY ASSURANCE ANALYSTS AND TESTERS"),      # -> Software Quality Assurance Analysts and Testers
    ("E018", "Rodrigo Campos",     "Finance",      "FINANCIAL AND INVESTMENT ANALYSTS"),                     # -> Financial and Investment Analysts
    ("E019", "Lucia Vargas",       "HR",           "HUMAN RESOURCES MANAGERS"),                             # -> Human Resources Managers
    ("E020", "Martin Alvarez",     "Data",         "DATA SCIENTISTS"),                                      # -> Data Scientists
    ("E021", "Claudia Reyes",      "IT",           "database administrators"),                              # -> Database Administrators
    ("E022", "Esteban Gutierrez",  "Security",     "information security analysts"),                        # -> Information Security Analysts
    ("E023", "Daniela Espinoza",   "Web",          "web developers"),                                       # -> Web Developers
    ("E024", "Javier Molina",      "Marketing",    "marketing managers"),                                   # -> Marketing Managers
    ("E025", "Isabela Mendoza",    "Ops",          "operations research analysts"),                         # -> Operations Research Analysts
    ("E026", "Camilo Ortega",      "Support",      "customer service representatives"),                     # -> Customer Service Representatives
    ("E027", "Paola Castillo",     "IT",           "COMPUTER-SYSTEMS ANALYSTS"),                            # -> Computer Systems Analysts
    ("E028", "Diego Fuentes",      "Nursing",      "REGISTERED NURSES"),                                    # -> Registered Nurses
    ("E029", "Valeria Herrera",    "Finance",      "ACCOUNTANTS AND AUDITORS"),                             # -> Accountants and Auditors
    ("E030", "Sebastian Rojas",    "Graphics",     "GRAPHIC DESIGNERS"),                                    # -> Graphic Designers
    # ── Type 3: Seniority prefix/suffix — corrected by pre_processor ─────────
    ("E031", "Alejandro Silva",    "Engineering",  "Senior Software Developer"),                            # -> Software Developers
    ("E032", "Carolina Mora",      "QA",           "Junior Software Quality Assurance Analyst"),            # -> Software Quality Assurance Analysts and Testers
    ("E033", "Mauricio Diaz",      "Data",         "Lead Data Scientist"),                                  # -> Data Scientists
    ("E034", "Xiomara Pena",       "Finance",      "Senior Financial and Investment Analyst"),              # -> Financial and Investment Analysts
    ("E035", "Emilio Castro",      "HR",           "Junior Human Resources Manager"),                      # -> Human Resources Managers
    ("E036", "Renata Vargas",      "IT",           "Senior Database Administrator"),                       # -> Database Administrators
    ("E037", "Oscar Rios",         "Security",     "Lead Information Security Analyst"),                   # -> Information Security Analysts
    ("E038", "Beatriz Torres",     "Web",          "Senior Web Developer"),                                # -> Web Developers
    ("E039", "Cristian Lopez",     "Marketing",    "Junior Marketing Manager"),                            # -> Marketing Managers
    ("E040", "Fernanda Gomez",     "Ops",          "Senior Operations Research Analyst"),                  # -> Operations Research Analysts
    ("E041", "Alvaro Munoz",       "Support",      "Lead Customer Service Representative"),                # -> Customer Service Representatives
    ("E042", "Manuela Ruiz",       "Engineering",  "Principal Software Developer"),                        # -> Software Developers
    ("E043", "Felipe Herrera",     "Network",      "Senior Network and Computer Systems Administrator"),   # -> Network and Computer Systems Administrators
    ("E044", "Veronica Mendez",    "Nursing",      "Lead Registered Nurse"),                               # -> Registered Nurses
    ("E045", "Nicolas Bravo",      "Graphics",     "Senior Graphic Designer"),                             # -> Graphic Designers
    # ── Type 4: Noise / context suffix — corrected by pre_processor ───────────
    ("E046", "Pamela Soto",        "Engineering",  "Software Developer - Remote (Contract)"),              # -> Software Developers
    ("E047", "Ignacio Vega",       "QA",           "Software Quality Assurance Analyst - Temp"),           # -> Software Quality Assurance Analysts and Testers
    ("E048", "Lorena Paredes",     "Finance",      "Financial and Investment Analyst (Freelance)"),        # -> Financial and Investment Analysts
    ("E049", "Eduardo Romero",     "HR",           "Human Resources Manager - Part Time"),                 # -> Human Resources Managers
    ("E050", "Gloria Navarro",     "Data",         "Data Scientist [Consultant]"),                         # -> Data Scientists
    ("E051", "Andres Delgado",     "IT",           "Database Administrator - On Site"),                    # -> Database Administrators
    ("E052", "Tatiana Guerrero",   "Security",     "Information Security Analyst (Remote)"),               # -> Information Security Analysts
    ("E053", "Rafael Cabrera",     "Web",          "Web Developer - Contract"),                            # -> Web Developers
    ("E054", "Adriana Pacheco",    "Marketing",    "Marketing Manager - Temp Position"),                   # -> Marketing Managers
    ("E055", "Hugo Jimenez",       "Support",      "Customer Service Representative (Part-time)"),         # -> Customer Service Representatives
    ("E056", "Sandra Leon",        "IT",           "Computer Systems Analyst - Remote Only"),              # -> Computer Systems Analysts
    ("E057", "Milton Flores",      "Nursing",      "Registered Nurse - Night Shift"),                      # -> Registered Nurses
    ("E058", "Karla Aguilar",      "Finance",      "Accountant and Auditor (External)"),                   # -> Accountants and Auditors
    ("E059", "Ernesto Blanco",     "Engineering",  "Software Developer - Contrato"),                       # -> Software Developers
    ("E060", "Pilar Ramirez",      "Graphics",     "Graphic Designer - Freelance"),                        # -> Graphic Designers
    # ── Type 5: Spanish translation — resolved by LLM ────────────────────────
    ("E061", "Rodrigo Montes",     "Engineering",  "Desarrollador de Software"),                           # -> Software Developers
    ("E062", "Diana Fuentes",      "QA",           "Analista de Calidad de Software"),                     # -> Software Quality Assurance Analysts and Testers
    ("E063", "Mauricio Serrano",   "Finance",      "Analista Financiero"),                                 # -> Financial and Investment Analysts
    ("E064", "Alejandra Castro",   "HR",           "Gerente de Recursos Humanos"),                         # -> Human Resources Managers
    ("E065", "Bruno Medina",       "Data",         "Cientifico de Datos"),                                 # -> Data Scientists
    ("E066", "Cecilia Torres",     "IT",           "Administrador de Bases de Datos"),                     # -> Database Administrators
    ("E067", "Damian Reyes",       "Security",     "Analista de Seguridad Informatica"),                   # -> Information Security Analysts
    ("E068", "Esperanza Ortiz",    "Web",          "Desarrollador Web"),                                   # -> Web Developers
    ("E069", "Francisco Leal",     "Marketing",    "Gerente de Marketing"),                                # -> Marketing Managers
    ("E070", "Helena Vargas",      "Ops",          "Analista de Investigacion Operativa"),                 # -> Operations Research Analysts
    ("E071", "Ivan Mendoza",       "Support",      "Representante de Atencion al Cliente"),                # -> Customer Service Representatives
    ("E072", "Jasmin Rios",        "IT",           "Analista de Sistemas Computacionales"),                # -> Computer Systems Analysts
    ("E073", "Kevin Salazar",      "Nursing",      "Enfermero Registrado"),                               # -> Registered Nurses
    ("E074", "Laura Ponce",        "Finance",      "Contador y Auditor"),                                  # -> Accountants and Auditors
    ("E075", "Marco Guzman",       "Graphics",     "Disenador Grafico"),                                   # -> Graphic Designers
    # ── Type 6: Abbreviation / synonym — resolved by LLM ─────────────────────
    ("E076", "Nadia Espinoza",     "HR",           "RRHH"),                                                # -> Human Resources Managers
    ("E077", "Omar Sanchez",       "Finance",      "CPA"),                                                 # -> Accountants and Auditors
    ("E078", "Patricia Molina",    "IT",           "DBA"),                                                 # -> Database Administrators
    ("E079", "Quintin Herrera",    "Engineering",  "Dev"),                                                 # -> Software Developers
    ("E080", "Raul Navarro",       "Security",     "SecOps"),                                              # -> Information Security Analysts
    ("E081", "Silvia Cabrera",     "Marketing",    "Mkt Manager"),                                         # -> Marketing Managers
    ("E082", "Tomas Guerrero",     "Web",          "Front-End Dev"),                                       # -> Web Developers
    ("E083", "Ursula Jimenez",     "Engineering",  "Back-End Dev"),                                        # -> Software Developers
    ("E084", "Vanessa Leon",       "Data",         "DS"),                                                  # -> Data Scientists
    ("E085", "Willian Flores",     "Support",      "CS Rep"),                                              # -> Customer Service Representatives
    ("E086", "Ximena Aguilar",     "IT",           "Sys Admin"),                                           # -> Network and Computer Systems Administrators
    ("E087", "Yahir Blanco",       "Engineering",  "Full Stack Dev"),                                      # -> Software Developers
    ("E088", "Zara Ramirez",       "Nursing",      "RN"),                                                  # -> Registered Nurses
    ("E089", "Alvaro Cruz",        "Finance",      "controller"),                                          # -> Financial Managers
    ("E090", "Beatriz Montes",     "HR",           "Talent Acquisition"),                                  # -> Human Resources Specialists
    # ── Type 7: Gender inflection (feminine Spanish) — resolved by LLM ───────
    ("E091", "Carmen Fuentes",     "Engineering",  "Desarrolladora de Software"),                          # -> Software Developers
    ("E092", "Daniela Serrano",    "QA",           "Analista de Calidad de Software"),                     # -> Software Quality Assurance Analysts and Testers
    ("E093", "Estela Castro",      "Finance",      "Analista Financiera"),                                 # -> Financial and Investment Analysts
    ("E094", "Francisca Medina",   "HR",           "Gerenta de Recursos Humanos"),                         # -> Human Resources Managers
    ("E095", "Gloria Torres",      "Data",         "Cientifica de Datos"),                                 # -> Data Scientists
    ("E096", "Helena Reyes",       "IT",           "Administradora de Bases de Datos"),                    # -> Database Administrators
    ("E097", "Iris Ortiz",         "Security",     "Analista de Seguridad Informatica"),                   # -> Information Security Analysts
    ("E098", "Jimena Leal",        "Web",          "Desarrolladora Web"),                                  # -> Web Developers
    ("E099", "Karina Vargas",      "Marketing",    "Gerenta de Marketing"),                                # -> Marketing Managers
    # 100th row — unrecognizable, goes to review queue (score < 0.70) ─────────
    ("E100", "Lucas Mendoza",      "Unknown",      "Xylophone Repair Technician"),                         # -> needs_review=True
]

HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _write_sheet(ws, column_name: str, id_col: str = "employee_id", name_col: str = "name", dept_col: str = "department") -> None:
    headers = [id_col, name_col, dept_col, column_name]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for row in ROWS:
        ws.append(list(row))

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 52


def _validate_canonicals(valid: set[str]) -> None:
    """Extract expected canonical titles from inline comments and verify against O*NET CSV."""
    import re
    with open(__file__) as f:
        source = f.read()

    # Parse lines like:  ("E001", ..., "anomaly"),   # -> Canonical Title
    pattern = re.compile(r'#\s*->\s*(.+)')
    errors: list[str] = []
    for line in source.splitlines():
        m = pattern.search(line)
        if not m:
            continue
        canonical = m.group(1).strip()
        if canonical in ("needs_review=True", "Canonical Title"):
            continue
        if canonical not in valid:
            errors.append(f"  NOT IN O*NET: {canonical!r}")

    if errors:
        print("VALIDATION ERRORS — fix before using as ground truth:")
        for e in errors:
            print(e)
    else:
        print(f"All canonical targets verified against O*NET ({len(valid)} titles). OK")


def main() -> None:
    valid = _load_valid()
    _validate_canonicals(valid)

    # File 1 — auto-detect (score >= 0.85 for "job_category")
    wb1 = openpyxl.Workbook()
    _write_sheet(wb1.active, "job_category")
    wb1.save("data/test_100_autodetect.xlsx")
    print("Created: data/test_100_autodetect.xlsx  (column: job_category -> auto-detect)")

    # File 2 — button fallback (ambiguous column names, all score < 0.85)
    wb2 = openpyxl.Workbook()
    _write_sheet(wb2.active, "headcount", id_col="emp_id", name_col="full_name", dept_col="dept_code")
    wb2.save("data/test_100_ambiguous.xlsx")
    print("Created: data/test_100_ambiguous.xlsx   (column: headcount -> button UI)")


if __name__ == "__main__":
    main()
