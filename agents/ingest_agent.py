import logging
import re

import openpyxl
from pydantic import BaseModel
from rapidfuzz import fuzz, process

from agno.workflow import OnError, Step, StepInput, StepOutput

from infrastructure.pipeline.session import PipelineSession
from infrastructure.pipeline.step_io import fail, ok

logger = logging.getLogger(__name__)

# Seed vocabulary for fuzzy header detection. Bilingual (EN + ES) because uploaded
# files come from mixed-language HR systems. No threshold is applied here — the
# best-matching column wins via WRatio regardless of language.
_JOB_COLUMN_SYNONYMS: list[str] = [
    # English — generic
    "job category", "job title", "job function", "job role", "job type",
    "position", "role", "occupation", "title", "function", "designation",
    "category", "profession", "trade", "job", "work category",
    # English — compound
    "job description", "job name", "job classification", "job code",
    "employment category", "staff category", "staff role", "work role",
    # Spanish — generic
    "cargo", "puesto", "categoria", "categoría", "ocupacion", "ocupación",
    "rol", "funcion", "función", "denominacion", "denominación", "empleo",
    # Spanish — compound
    "tipo cargo", "categoria empleo", "categoria laboral", "nombre cargo",
    "descripcion cargo", "descripcion puesto", "puesto trabajo",
    "tipo empleo", "perfil", "perfil laboral", "area funcional",
    "puesto de trabajo", "categoria de empleo", "categoria de trabajo",
]


# Intermediate result used by the UI to present column choices to the user before
# the target column is confirmed. Kept separate from IngestResult so the two-step
# flow (scan → select → extract) is explicit in the type system.
class HeaderScanResult(BaseModel):
    file_path: str
    column_names: list[str]


class IngestResult(BaseModel):
    """
    The output of extract_categories() — everything downstream steps need from the source file.

    raw_categories is deduplicated so MapperAgent processes each unique title once,
    not once per row. AuditWriter uses file_path + target_column to reopen the
    original workbook and write corrections back to the right column.
    """
    file_path: str
    target_column: str
    raw_categories: list[str]   # unique, sorted, stripped
    # total_rows includes blanks; len(raw_categories) does not — see extract_categories
    total_rows: int


def scan_headers(file_path: str) -> HeaderScanResult:
    """
    Read only the first row of the workbook and return the column names.

    This is intentionally a separate step from extract_categories so the UI can
    present column choices to the user (or run detect_job_column) before committing
    to a target column — avoiding a full file read if the user changes their mind.
    """
    wb = openpyxl.load_workbook(file_path, read_only=True)
    ws = wb.active
    # Filter None to skip trailing empty columns that openpyxl returns when the
    # sheet has phantom cells beyond the actual data range.
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1)) if cell.value is not None]
    # Explicit close is required for read_only workbooks — they hold a file handle
    # that is not released until close() or GC, whichever comes last.
    wb.close()
    return HeaderScanResult(file_path=file_path, column_names=headers)


def detect_job_column(column_names: list[str]) -> tuple[str | None, float]:
    """Returns (best_column_name, score 0.0-1.0). (None, 0.0) when list is empty."""
    if not column_names:
        return None, 0.0

    best_col: str | None = None
    best_score: float = 0.0

    for col in column_names:
        # Collapse separators before matching so "job_title", "job-title", and
        # "job title" all produce identical tokens and score the same.
        normalized = re.sub(r"[\s_\-\.]+", " ", col.lower()).strip()
        # WRatio handles length mismatches better than simple ratio; preferred for
        # short strings like column headers where token order may vary.
        match = process.extractOne(normalized, _JOB_COLUMN_SYNONYMS, scorer=fuzz.WRatio)
        if match:
            # rapidfuzz returns 0-100; normalise to 0.0-1.0 to match the
            # pipeline-wide confidence scale used in the confidence scoring layer.
            score = round(match[1] / 100.0, 4)
            if score > best_score:
                best_col = col
                best_score = score

    return best_col, best_score


def extract_categories(file_path: str, target_column: str) -> IngestResult:
    """
    Read every data row and collect the unique, non-blank values in target_column.

    Deduplication happens here (not in ValidatorAgent) so the LLM-path agents
    process each distinct title exactly once regardless of how many rows contain it.
    total_rows counts ALL data rows including blanks so AuditWriter can report
    coverage accurately even when some rows had no value in the target column.
    """
    # read_only=True uses openpyxl's optimised read path — avoids loading cell
    # styles and formula cache, which matters for large HR exports (>10k rows).
    wb = openpyxl.load_workbook(file_path, read_only=True)
    ws = wb.active

    header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    if target_column not in header_row:
        raise ValueError(f"Column '{target_column}' not found. Available: {header_row}")

    col_idx = header_row.index(target_column)

    raw_values: list[str] = []
    total_rows = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        total_rows += 1  # counts ALL data rows, including blanks in the target column
        val = row[col_idx]
        if val is not None:
            raw_values.append(str(val).strip())

    wb.close()

    # Deduplicate and sort so downstream agents receive a stable, predictable list.
    unique_categories = sorted(set(raw_values))
    return IngestResult(
        file_path=file_path,
        target_column=target_column,
        raw_categories=unique_categories,
        total_rows=total_rows,
    )


def ingest_executor(_step_input: StepInput, session_state: dict) -> StepOutput:
    # Agno passes shared pipeline state as a plain dict; PipelineSession.from_dict
    # deserialises it into a typed object with file_path and target_column.
    try:
        session = PipelineSession.from_dict(session_state)
        result = extract_categories(session.file_path, session.target_column)
        logger.info(
            "ingest: column=%r total_rows=%d unique_categories=%d",
            result.target_column,
            result.total_rows,
            len(result.raw_categories),
        )
        # ok() wraps the result in a successful StepOutput for the next step.
        return ok(result)
    except Exception as e:
        # fail() wraps the exception and signals the Workflow to stop this step.
        return fail(e)


# on_error=fail halts the entire Workflow immediately; there is no meaningful
# recovery if the source file is unreadable or the target column is missing.
ingest_step = Step(name="ingest", executor=ingest_executor, on_error=OnError.fail)