"""
Writes the corrected Excel file and audit log after all mapping decisions are finalized.
Called by the Workflow as the last step, after MapperAgent has produced a MappingResult.
Hard invariant: a correction is ONLY written if it is an exact O*NET title — any LLM
hallucination is caught here and routed to the review queue instead.
"""
import logging
from datetime import datetime
from pathlib import Path

import openpyxl

from pydantic import BaseModel

from agno.workflow import OnError, Step, StepInput, StepOutput

from agents.mapper_agent import MappingDecision, MappingResult
from domain.onet import is_valid_onet_title
from infrastructure.pipeline.session import PipelineSession
from infrastructure.pipeline.step_io import deserialize, fail, ok

logger = logging.getLogger(__name__)


class AuditResult(BaseModel):
    """Summary statistics produced after writing the output Excel file."""

    output_path: str          # Absolute file path of the corrected Excel workbook.
    corrected_count: int      # Number of rows where a correction was successfully applied.
    review_queue_count: int   # Number of rows sent to the human review queue (not auto-corrected).
    hallucination_count: int  # Number of LLM suggestions rejected for naming a non-existent O*NET title.
    precision: float | None   # corrected / (corrected + hallucination); None if no corrections attempted


def _write_excel(
    source_path: str,
    decisions_by_raw: dict[str, MappingDecision],
    target_column: str,
    valid_categories_set: set[str],
) -> tuple[str, int, int, int]:
    """
    Build and save the two-sheet output workbook, then return outcome counts.

    The output file contains two sheets:
      - "Corrected": every original row, plus three new columns showing what
        correction was applied (or left blank if the row needs human review).
      - "Review Queue": only the rows that could not be corrected automatically,
        with a short reason explaining why each one needs a human decision.

    This function also acts as the final hallucination guard: even if MapperAgent
    returned a correction, it is verified here against valid_categories_set before
    being written. Any title that is not an exact O*NET canonical title is rejected
    and routed to the Review Queue instead.
    """
    wb_in = openpyxl.load_workbook(source_path)
    ws_in = wb_in.active

    headers = [cell.value for cell in next(ws_in.iter_rows(min_row=1, max_row=1))]
    col_idx = headers.index(target_column)

    wb_out = openpyxl.Workbook()

    # Sheet 1: all original rows with correction columns appended.
    ws_corrected = wb_out.active
    ws_corrected.title = "Corrected"
    ws_corrected.append(headers + ["corrected_category", "correction_method", "confidence"])

    # Sheet 2: the audit log — rows a human must inspect before accepting.
    ws_review = wb_out.create_sheet("Review Queue")
    ws_review.append(["original_category", "preprocessed", "review_reason"])

    corrected_count = 0
    review_queue_count = 0
    hallucination_count = 0

    for row in ws_in.iter_rows(min_row=2, values_only=True):
        raw_val = str(row[col_idx]).strip() if row[col_idx] is not None else None
        decision = decisions_by_raw.get(raw_val) if raw_val else None

        if decision is None:
            # Category was already valid — pass through unchanged
            # (exact O*NET title detected at the ValidatorAgent step).
            ws_corrected.append(list(row) + [raw_val, "exact", 1.0])
            continue

        if decision.needs_review:
            # 'needs_review' means the pipeline's confidence was too low to auto-correct
            # (rapidfuzz score < 0.70). The system never guesses in this band.
            # The correction column is left blank; the row goes to the Review Queue sheet.
            ws_corrected.append(list(row) + [None, decision.method, decision.confidence])
            ws_review.append([decision.raw, decision.preprocessed, decision.review_reason or "needs_review"])
            review_queue_count += 1
            continue

        corrected = decision.corrected
        # Hard invariant: verify correction is a valid O*NET title before writing.
        # This is the hallucination guard — it catches cases where the LLM returned
        # a title that sounds plausible but does not exist in valid_categories.csv.
        if not is_valid_onet_title(corrected, valid_categories_set):
            # Reject the hallucination: blank the correction and send to review queue.
            ws_corrected.append(list(row) + [None, "needs_review", decision.confidence])
            ws_review.append([decision.raw, decision.preprocessed, "hallucination_rejected"])
            hallucination_count += 1
            review_queue_count += 1
            continue

        ws_corrected.append(list(row) + [corrected, decision.method, decision.confidence])
        corrected_count += 1

    wb_in.close()

    # Build a timestamped output filename to avoid overwriting previous runs.
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    source_stem = Path(source_path).stem
    output_dir = Path(source_path).parent
    output_path = str(output_dir / f"{source_stem}_corrected_{timestamp}.xlsx")
    wb_out.save(output_path)

    return output_path, corrected_count, review_queue_count, hallucination_count


def audit_executor(step_input: StepInput, session_state: dict) -> StepOutput:
    """
    Agno Step executor that drives the full audit-and-write cycle.

    Pipeline order: IngestAgent -> ValidatorAgent -> MapperAgent -> AuditWriter (this step)
    """
    try:
        session = PipelineSession.from_dict(session_state)
        mapping_result = deserialize(step_input.previous_step_content, MappingResult)

        # Build a lookup dict so each row in the Excel file can find its decision in O(1).
        decisions_by_raw = {d.raw: d for d in mapping_result.decisions}

        output_path, corrected_count, review_queue_count, hallucination_count = _write_excel(
            source_path=session.file_path,
            decisions_by_raw=decisions_by_raw,
            target_column=session.target_column,
            valid_categories_set=session.valid_categories_set,
        )

        # Precision = share of auto-corrections that were actually correct.
        # None when nothing was attempted (e.g. all rows were already valid O*NET titles).
        total_attempted = corrected_count + hallucination_count
        precision = corrected_count / total_attempted if total_attempted > 0 else None

        result = AuditResult(
            output_path=output_path,
            corrected_count=corrected_count,
            review_queue_count=review_queue_count,
            hallucination_count=hallucination_count,
            precision=precision,
        )
        logger.info(
            "audit: corrected=%d review=%d hallucinations=%d precision=%s output=%r",
            corrected_count,
            review_queue_count,
            hallucination_count,
            f"{precision:.4f}" if precision is not None else "N/A",
            output_path,
        )
        return ok(result)
    except Exception as e:
        return fail(e)


# Register this executor as a named Agno Step.
# on_error=OnError.fail stops the entire Workflow if this step raises —
# an incomplete audit must never be silently swallowed.
audit_step = Step(name="audit", executor=audit_executor, on_error=OnError.fail)